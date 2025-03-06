from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import RED, GREEN
from datetime import datetime
import json

app = Flask(__name__)
print("Hello World")

name = ""

@app.route('/', methods = ['POST', 'GET'])
def login():
    with open('authorized.json') as json_file:
        data = json.load(json_file)
    if request.method == 'GET':
        return render_template('login.html')
    else:
        if request.form['madrasahID'].upper() in data.keys():
            name = data[request.form['madrasahID'].upper()]["Name"]
            return render_template('mainPg.html', name = name, classList = data[request.form['madrasahID'].upper()]["Class"])
            #print(data[request.form['madrasahID'].upper()]["Class"])

        else:
            return "Login Error"


@app.route('/attendance/<className>', methods = ['POST', 'GET'])
def attendPg(className):
    listNames, className = getNames(className)
    if request.method == 'GET':
        return render_template('attndPg.html', name = name, namesList = listNames, className = className)
    #else:
        #return updateExcel(request.form.to_dict(), wb, ws)

@app.route('/download')
def downloadFile ():
    path = "C:/Users/mriswan/Desktop/ProjX/test.xlsx"
    return send_file(path, as_attachment=True)

def getNames(className):
    namesList = []
    wb = load_workbook('YBH ATTENDANCE LIST 2019.xlsx')
    ws = wb[className]
    for row in ws.iter_rows(min_row=6, max_col=3, values_only=True):
        if(row[2] is None):
            break
        else:
            namesList.append(row)
    return namesList, className

@app.route('/update/<className>', methods = ['POST', 'GET'])
def updateExcel(className):
    attendData = request.form.to_dict()
    wb = load_workbook('YBH ATTENDANCE LIST 2019.xlsx')
    ws = wb[className]
    currDate = datetime.now()
    today = datetime(currDate.year, currDate.month, currDate.day)
    dateCol = []
    for row in ws.iter_rows(min_col=4, min_row=5, max_row = 5, values_only=True):
        for value in row:
            if(value is None):
                break
            else:
                dateCol.append(value)

    #print (dateCol)
    if today not in dateCol:
        dateCol.append(today)

    colNum = dateCol.index(today)
    print(colNum)
    cell = ws.cell(row = 5, column = 4+colNum)
    cell.value = today
    presentFt = Font(color=GREEN)
    absentFt = Font(color=RED)
    
    for i in range(6, 6+len(attendData)):
        nameCell = ws.cell(row = i, column = 3)
        if nameCell.value in attendData:
            buffCell = ws.cell(row = i, column = 4+colNum)
            buffCell.value = int(attendData[nameCell.value])
            if buffCell.value == 1:
                buffCell.font = presentFt
            elif buffCell.value == 0:
                buffCell.font = absentFt

    wb.save('test.xlsx')
    return 'updated'

if __name__ == '__main__':
   app.run(debug = True)