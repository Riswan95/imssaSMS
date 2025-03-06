from openpyxl import load_workbook, Workbook
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import Column, Integer, String, Enum, Float, ForeignKey, Date, Table
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker, relationship

engine = create_engine('sqlite:///ybh_imssa1.db')
base = declarative_base()

#teacher_class_table (many to many)
teacher_class = Table('teacher_class_table', base.metadata,
Column('id', Integer, primary_key=True), 
Column('teacher_id', String, ForeignKey('Teacher.id')),
Column('class_name', String, ForeignKey('Class.class_name')))

#Class Table
class Class(base):
    __tablename__ = 'Class'
    
    #variables/Column Names
    class_name = Column(String, primary_key=True)

    #relationships
    students = relationship("Student", back_populates="Class")
    teachers = relationship("Teacher", secondary= teacher_class, back_populates="classes")

#Student Table
class Student(base):
    __tablename__ = 'Student'
    
    #variables/Column Names
    student_id = Column(String, primary_key=True)
    student_name = Column(String)
    class_name = Column(String, ForeignKey('Class.class_name'))

    #relationships
    Class = relationship("Class", back_populates = "students")
    attendances = relationship('Attendance', back_populates='student')

#Class Attendance
class Attendance(base):
    __tablename__ = 'Attendance'
    
    #variables/Column Names
    id = Column(Integer, primary_key=True)
    date = Column(Date)
    presence = Column(Integer)
    student_id = Column(String, ForeignKey('Student.student_id'))

    #relationships
    student = relationship("Student", back_populates = "attendances")

#Teacher Table
class Teacher(base):
    __tablename__ = 'Teacher'
    
    #variables/Column Names
    id = Column(String, primary_key=True)
    #password = Column(String(80))

    #relationships
    classes = relationship("Class", secondary= teacher_class, back_populates='teachers')


base.metadata.create_all(engine) #creates the database
Session = sessionmaker(bind=engine)
session = Session()

def getNames(className):    
    namesList = []
    wb = load_workbook('YBH ATTENDANCE LIST 2019.xlsx')
    ws = wb[className]
    for row in ws.iter_rows(min_row=6, max_col=3, values_only=True):
        if(row[2] is None):
            break
        else:
            namesList.append(row)
    return namesList

def addStudents2class():
    #buffer_class = Class(class_name = (input('Which class (in DB) would you like to add these students?: ')).upper())
    buffer_class = session.query(Class).filter(Class.class_name == (input('Which class (in DB) would you like to add these students?: ')).upper()).first()
    if buffer_class is None:
        print('Class does not exist in db')
        #addClass()
    else:
        className = (input('Enter Class Name as in Excel Sheet: ')).upper()
        namesList = getNames(className)
        for data in namesList:
            buffer_student = Student(student_id = data[1], student_name = data[2])
            buffer_class.students.append(buffer_student)
            session.add(buffer_student)
            print(buffer_student.student_name, 'added!')
        session.commit()
        session.close()
        print("added all students to db!")


def addClass2db():
    className = (input('Enter Class Name: ')).upper()
    buffer_class = Class(class_name = className)
    session.add(buffer_class)
    session.commit()
    session.close()
    print("class added to db!")

def addTeacher2db():
    buffer_student = session.query(Student).filter(Student.student_id == (input('Enter ID of Teacher: ').upper())).first()
    buffer_teacher = Teacher(id=buffer_student.student_id)
    session.add(buffer_teacher)
    session.commit()
    session.close()
    print("teacher added to db!")

def assignTeacher2class():
    buffer_class = session.query(Class).filter(Class.class_name == (input('Which class (in DB) would you like to assign a teacher?: ')).upper()).first()
    if buffer_class is None:
        print('Class does not exist in db')
        #addClass()
    else:
       buffer_teacher = session.query(Teacher).filter(Teacher.id == (input('Whats the teachers id?: ')).upper()).first()
       buffer_class.teachers.append(buffer_teacher)
       session.add(buffer_class)
       session.commit()
       session.close()

addClass2db()
addStudents2class()
addTeacher2db()
assignTeacher2class()